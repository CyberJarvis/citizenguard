"""
Export Service
Handles CSV, Excel, and PDF export generation for the Analyst module.
"""

import logging
import os
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Any
from pathlib import Path
from io import BytesIO

from motor.motor_asyncio import AsyncIOMotorDatabase

from app.models.analyst import (
    ExportJob, ExportStatus, ExportFormat, ExportType
)
from app.middleware.rbac import Permission
from app.models.rbac import RolePermissions, UserRole

logger = logging.getLogger(__name__)

# PII fields that must be removed from exports
PII_FIELDS = [
    'name', 'email', 'phone', 'address', 'full_name', 'user_name',
    'contact', 'personal_info', 'profile_picture', 'reporter_name',
    'user_email', 'user_phone'
]

# Export directory
EXPORT_DIR = Path("exports")
EXPORT_DIR.mkdir(exist_ok=True)


class ExportService:
    """
    Service for generating data exports in various formats.

    Supports:
    - CSV: Simple tabular data
    - Excel: Formatted spreadsheets with multiple sheets
    - PDF: Formatted reports with charts

    All exports automatically filter PII for analyst users.
    """

    def __init__(self, db: AsyncIOMotorDatabase):
        self.db = db

    def sanitize_pii(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Remove PII fields from data before export.

        Args:
            data: Dictionary containing report/user data

        Returns:
            Sanitized dictionary with PII removed
        """
        sanitized = {}
        for key, value in data.items():
            if key.lower() not in [f.lower() for f in PII_FIELDS]:
                if isinstance(value, dict):
                    sanitized[key] = self.sanitize_pii(value)
                elif isinstance(value, list):
                    sanitized[key] = [
                        self.sanitize_pii(item) if isinstance(item, dict) else item
                        for item in value
                    ]
                else:
                    sanitized[key] = value

        # Add anonymized reporter info
        if 'user_id' in data:
            sanitized['reporter_id'] = data['user_id']
            sanitized['reporter_type'] = 'VERIFIED_CITIZEN'

        return sanitized

    def _clean_query_filters(self, query_config: Dict[str, Any]) -> Dict[str, Any]:
        """
        Clean query config by removing:
        - Internal keys (starting with _)
        - UI placeholder values ('all', 'All', '', None)
        - Pagination/sort keys
        """
        if not query_config:
            return {}

        exclude_values = {'all', 'All', '', None}
        exclude_keys = {'page', 'limit', 'skip', 'sort', 'sort_by', 'sort_order'}

        cleaned = {}
        for key, value in query_config.items():
            # Skip internal keys
            if key.startswith('_'):
                continue
            # Skip pagination/sort keys
            if key in exclude_keys:
                continue
            # Skip placeholder values
            if value in exclude_values:
                continue
            # Skip empty lists
            if isinstance(value, list) and len(value) == 0:
                continue
            # Include valid values
            cleaned[key] = value

        return cleaned

    async def get_export_data(
        self,
        export_type: ExportType,
        query_config: Dict[str, Any],
        date_range: Dict[str, Any],
        columns: Optional[List[str]] = None
    ) -> List[Dict[str, Any]]:
        """
        Fetch data for export based on configuration.

        Args:
            export_type: Type of data to export
            query_config: Query filters
            date_range: Date range for data
            columns: Optional list of columns to include

        Returns:
            List of sanitized data records
        """
        # Clean query config - remove internal keys and placeholder values like 'all'
        clean_query_config = self._clean_query_filters(query_config)

        logger.info(f"get_export_data called: type={export_type}, date_range={date_range}, filters={clean_query_config}")

        # Build date filter
        date_filter = {}
        if date_range and date_range.get("start"):
            try:
                start_str = date_range["start"]
                if isinstance(start_str, str):
                    start_str = start_str.replace('Z', '+00:00')
                    date_filter["created_at"] = {"$gte": datetime.fromisoformat(start_str)}
            except (ValueError, AttributeError) as e:
                logger.warning(f"Failed to parse start date: {e}")

        if date_range and date_range.get("end"):
            try:
                end_str = date_range["end"]
                if isinstance(end_str, str):
                    end_str = end_str.replace('Z', '+00:00')
                    if "created_at" in date_filter:
                        date_filter["created_at"]["$lte"] = datetime.fromisoformat(end_str)
                    else:
                        date_filter["created_at"] = {"$lte": datetime.fromisoformat(end_str)}
            except (ValueError, AttributeError) as e:
                logger.warning(f"Failed to parse end date: {e}")

        # Handle relative date ranges
        if date_range and date_range.get("relative"):
            now = datetime.now(timezone.utc)
            relative = date_range["relative"]
            if relative == "7days":
                start = now - timedelta(days=7)
            elif relative == "30days":
                start = now - timedelta(days=30)
            elif relative == "90days":
                start = now - timedelta(days=90)
            elif relative == "year":
                start = now - timedelta(days=365)
            elif relative == "all":
                # No date filter for "all time"
                date_filter = {}
                start = None
            else:
                start = now - timedelta(days=7)

            if relative != "all":
                date_filter["created_at"] = {"$gte": start, "$lte": now}

        # Determine collection and query based on export type
        if export_type == ExportType.REPORTS:
            collection = self.db.hazard_reports
            query = {**date_filter, **clean_query_config}
            logger.info(f"Reports query: {query}")

            # Fetch data
            cursor = collection.find(query).sort("created_at", -1).limit(10000)
            data = await cursor.to_list(10000)
            logger.info(f"Reports found: {len(data)} records")

            # Sanitize PII
            sanitized_data = [self.sanitize_pii(doc) for doc in data]

            # Convert ObjectId to string and handle datetime
            for record in sanitized_data:
                if "_id" in record:
                    record["_id"] = str(record["_id"])
                for key, value in list(record.items()):
                    if isinstance(value, datetime):
                        record[key] = value.isoformat()

            logger.info(f"Returning {len(sanitized_data)} sanitized report records")
            return sanitized_data

        elif export_type == ExportType.ANALYTICS:
            # Return aggregated analytics data
            from app.services.analytics_service import get_analytics_service
            analytics = get_analytics_service(self.db)

            relative = date_range.get("relative", "30days")
            report_analytics = await analytics.get_report_analytics(date_range=relative)
            hazard_analytics = await analytics.get_hazard_type_analytics(date_range=relative)

            # Build metrics from report_analytics (keys are at top level, not under "metrics")
            return [{
                "type": "analytics_summary",
                "report_metrics": {
                    "total_reports": report_analytics.get("total_reports", 0),
                    "verified_reports": report_analytics.get("verified_reports", 0),
                    "pending_reports": report_analytics.get("pending_reports", 0),
                    "high_severity_count": report_analytics.get("high_severity_count", 0),
                    "verified_percentage": round(
                        (report_analytics.get("verified_reports", 0) / max(report_analytics.get("total_reports", 1), 1)) * 100, 1
                    )
                },
                "by_status": report_analytics.get("by_status", {}),
                "by_hazard_type": report_analytics.get("by_hazard_type", {}),
                "by_region": report_analytics.get("by_region", {}),
                "hazard_details": hazard_analytics.get("hazard_types", [])
            }]

        elif export_type == ExportType.TRENDS:
            from app.services.analytics_service import get_analytics_service
            analytics = get_analytics_service(self.db)

            relative = date_range.get("relative", "30days") if date_range else "30days"
            group_by = clean_query_config.get("group_by", "day") if clean_query_config else "day"
            trend_data = await analytics.get_trend_data(
                date_range=relative,
                group_by=group_by
            )

            timeline = trend_data.get("timeline", [])
            logger.info(f"Trends data: {len(timeline)} data points")
            # Return empty placeholder if no data
            if not timeline:
                return [{"date": "No data", "total": 0, "verified": 0, "pending": 0, "high_priority": 0}]
            return timeline

        elif export_type == ExportType.GEO:
            from app.services.analytics_service import get_analytics_service
            analytics = get_analytics_service(self.db)

            relative = date_range.get("relative", "30days") if date_range else "30days"
            geo_data = await analytics.get_geospatial_data(date_range=relative)

            # Return regions with counts
            regions = geo_data.get("regions", [])
            logger.info(f"Geo data: {len(regions)} regions")

            # If no regions, try to get report_points as fallback
            if not regions:
                report_points = geo_data.get("report_points", [])
                if report_points:
                    return report_points
                return [{"name": "No data", "count": 0, "center": {"lat": 0, "lng": 0}}]
            return regions

        else:
            # Custom - fetch reports by default
            collection = self.db.hazard_reports
            query = {**date_filter, **clean_query_config}
            logger.info(f"Custom query: {query}")
            cursor = collection.find(query).sort("created_at", -1).limit(10000)
            data = await cursor.to_list(10000)
            logger.info(f"Custom data: {len(data)} records")
            return [self.sanitize_pii(doc) for doc in data]

    async def generate_csv(
        self,
        data: List[Dict[str, Any]],
        columns: Optional[List[str]] = None
    ) -> bytes:
        """
        Generate CSV file from data.

        Args:
            data: List of records
            columns: Optional list of columns to include

        Returns:
            CSV file content as bytes
        """
        import pandas as pd

        if not data:
            return b"No data available"

        # Create DataFrame
        df = pd.DataFrame(data)

        # Filter columns if specified
        if columns:
            available_columns = [c for c in columns if c in df.columns]
            if available_columns:
                df = df[available_columns]

        # Flatten nested dictionaries
        for col in df.columns:
            if df[col].apply(lambda x: isinstance(x, dict)).any():
                # Convert dict columns to JSON strings
                df[col] = df[col].apply(
                    lambda x: str(x) if isinstance(x, dict) else x
                )

        # Generate CSV
        output = BytesIO()
        df.to_csv(output, index=False, encoding='utf-8')
        return output.getvalue()

    async def generate_excel(
        self,
        data: List[Dict[str, Any]],
        export_type: ExportType,
        columns: Optional[List[str]] = None
    ) -> bytes:
        """
        Generate Excel file from data.

        Args:
            data: List of records
            export_type: Type of export for sheet naming
            columns: Optional list of columns to include

        Returns:
            Excel file content as bytes
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows

        if not data:
            # Return empty workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "No Data"
            ws['A1'] = "No data available for the selected criteria"
            output = BytesIO()
            wb.save(output)
            return output.getvalue()

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = export_type.value.capitalize()

        # Create DataFrame
        df = pd.DataFrame(data)

        # Filter columns if specified
        if columns:
            available_columns = [c for c in columns if c in df.columns]
            if available_columns:
                df = df[available_columns]

        # Flatten nested dictionaries
        for col in df.columns:
            if df[col].apply(lambda x: isinstance(x, dict)).any():
                df[col] = df[col].apply(
                    lambda x: str(x) if isinstance(x, dict) else x
                )

        # Write data
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Style header row
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to bytes
        output = BytesIO()
        wb.save(output)
        return output.getvalue()

    async def generate_pdf(
        self,
        data: List[Dict[str, Any]],
        export_type: ExportType,
        title: str = "CoastGuardian Analytics Report"
    ) -> bytes:
        """
        Generate PDF report from data.

        Args:
            data: List of records or analytics data
            export_type: Type of export for formatting
            title: Report title

        Returns:
            PDF file content as bytes
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT

        output = BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )

        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=30
        )
        story.append(Paragraph(title, title_style))

        # Generation info
        info_style = ParagraphStyle(
            'Info',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=colors.grey
        )
        story.append(Paragraph(
            f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} UTC",
            info_style
        ))
        story.append(Spacer(1, 20))

        if not data:
            story.append(Paragraph("No data available for the selected criteria.", styles['Normal']))
        elif export_type == ExportType.ANALYTICS:
            # Analytics summary report
            analytics_data = data[0] if data else {}

            # Metrics section
            story.append(Paragraph("Report Metrics", styles['Heading2']))
            story.append(Spacer(1, 10))

            metrics = analytics_data.get('report_metrics', {})
            metrics_data = [
                ['Metric', 'Value'],
                ['Total Reports', str(metrics.get('total_reports', 0))],
                ['Verified Reports', str(metrics.get('verified_reports', 0))],
                ['Pending Reports', str(metrics.get('pending_reports', 0))],
                ['High Severity', str(metrics.get('high_severity_count', 0))],
                ['Verification Rate', f"{metrics.get('verified_percentage', 0)}%"],
            ]

            t = Table(metrics_data, colWidths=[3*inch, 2*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            story.append(t)
            story.append(Spacer(1, 20))

            # Hazard types section
            hazard_types = analytics_data.get('by_hazard_type', {})
            if hazard_types:
                story.append(Paragraph("Reports by Hazard Type", styles['Heading2']))
                story.append(Spacer(1, 10))

                hazard_data = [['Hazard Type', 'Count']]
                for hazard, count in sorted(hazard_types.items(), key=lambda x: x[1], reverse=True)[:10]:
                    hazard_data.append([hazard.replace('_', ' ').title(), str(count)])

                t = Table(hazard_data, colWidths=[3.5*inch, 1.5*inch])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ]))
                story.append(t)

        elif export_type in [ExportType.REPORTS, ExportType.CUSTOM]:
            # Table of reports
            story.append(Paragraph(f"Hazard Reports ({len(data)} records)", styles['Heading2']))
            story.append(Spacer(1, 10))

            # Select key columns for PDF
            key_columns = ['report_id', 'hazard_type', 'verification_status', 'risk_level', 'created_at']
            headers = ['Report ID', 'Hazard Type', 'Status', 'Risk Level', 'Created At']

            table_data = [headers]
            for record in data[:50]:  # Limit to 50 for PDF
                row = []
                for col in key_columns:
                    value = record.get(col, '')
                    if col == 'hazard_type':
                        value = str(value).replace('_', ' ').title() if value else ''
                    elif col == 'created_at':
                        value = str(value)[:10] if value else ''
                    else:
                        value = str(value)[:20] if value else ''
                    row.append(value)
                table_data.append(row)

            if len(data) > 50:
                table_data.append(['...', f'{len(data) - 50} more records', '', '', ''])

            t = Table(table_data, colWidths=[1.2*inch, 1.2*inch, 1*inch, 0.9*inch, 1*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#E8E8E8')]),
            ]))
            story.append(t)

        elif export_type == ExportType.TRENDS:
            # Trend data
            story.append(Paragraph(f"Trend Analysis ({len(data)} data points)", styles['Heading2']))
            story.append(Spacer(1, 10))

            table_data = [['Date', 'Total', 'Verified', 'Pending', 'High Priority']]
            for record in data:
                table_data.append([
                    str(record.get('date', '')),
                    str(record.get('total', 0)),
                    str(record.get('verified', 0)),
                    str(record.get('pending', 0)),
                    str(record.get('high_priority', 0)),
                ])

            t = Table(table_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(t)

        elif export_type == ExportType.GEO:
            # Geographic data
            story.append(Paragraph(f"Geographic Distribution ({len(data)} regions)", styles['Heading2']))
            story.append(Spacer(1, 10))

            table_data = [['Region', 'Report Count', 'Center Lat', 'Center Lon']]
            for record in data:
                center = record.get('center', {})
                # Support both 'lng' (from analytics service) and 'lon' formats
                longitude = center.get('lng', center.get('lon', 0))
                table_data.append([
                    str(record.get('name', '')),
                    str(record.get('count', 0)),
                    f"{center.get('lat', 0):.4f}",
                    f"{longitude:.4f}",
                ])

            t = Table(table_data, colWidths=[2.5*inch, 1.2*inch, 1.2*inch, 1.2*inch])
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ]))
            story.append(t)

        # Footer
        story.append(Spacer(1, 30))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.grey
        )
        story.append(Paragraph(
            "CoastGuardian - Ocean Hazard Reporting Platform for INCOIS",
            footer_style
        ))
        story.append(Paragraph(
            "This report was generated for authorized analysts only. PII has been filtered.",
            footer_style
        ))

        # Build PDF
        doc.build(story)
        return output.getvalue()

    async def process_export_job(self, job_id: str) -> bool:
        """
        Process an export job and generate the file.

        Args:
            job_id: Export job ID

        Returns:
            True if successful, False otherwise
        """
        try:
            # Get job from database
            job_doc = await self.db.export_jobs.find_one({"job_id": job_id})
            if not job_doc:
                logger.error(f"Export job {job_id} not found")
                return False

            job = ExportJob.from_mongo(job_doc)

            # Update status to processing
            await self.db.export_jobs.update_one(
                {"job_id": job_id},
                {
                    "$set": {
                        "status": ExportStatus.PROCESSING.value,
                        "started_at": datetime.now(timezone.utc),
                        "progress": 10
                    }
                }
            )

            # Fetch data
            await self.db.export_jobs.update_one(
                {"job_id": job_id},
                {"$set": {"progress": 30}}
            )

            data = await self.get_export_data(
                export_type=job.export_type,
                query_config=job.query_config,
                date_range=job.date_range,
                columns=job.columns
            )

            await self.db.export_jobs.update_one(
                {"job_id": job_id},
                {"$set": {"progress": 60}}
            )

            # Generate file based on format
            if job.export_format == ExportFormat.CSV:
                content = await self.generate_csv(data, job.columns)
                extension = "csv"
                content_type = "text/csv"
            elif job.export_format == ExportFormat.EXCEL:
                content = await self.generate_excel(data, job.export_type, job.columns)
                extension = "xlsx"
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            elif job.export_format == ExportFormat.PDF:
                content = await self.generate_pdf(data, job.export_type)
                extension = "pdf"
                content_type = "application/pdf"
            else:
                raise ValueError(f"Unsupported export format: {job.export_format}")

            await self.db.export_jobs.update_one(
                {"job_id": job_id},
                {"$set": {"progress": 90}}
            )

            # Save file
            file_name = f"{job_id}_{job.export_type.value}_{datetime.now().strftime('%Y%m%d')}.{extension}"
            file_path = EXPORT_DIR / file_name

            with open(file_path, 'wb') as f:
                f.write(content)

            # Update job as completed
            await self.db.export_jobs.update_one(
                {"job_id": job_id},
                {
                    "$set": {
                        "status": ExportStatus.COMPLETED.value,
                        "progress": 100,
                        "completed_at": datetime.now(timezone.utc),
                        "file_path": str(file_path),
                        "file_name": file_name,
                        "file_size": len(content),
                        "record_count": len(data) if isinstance(data, list) else 1
                    }
                }
            )

            logger.info(f"Export job {job_id} completed successfully: {file_name}")
            return True

        except Exception as e:
            logger.error(f"Error processing export job {job_id}: {e}")

            # Update job as failed
            await self.db.export_jobs.update_one(
                {"job_id": job_id},
                {
                    "$set": {
                        "status": ExportStatus.FAILED.value,
                        "error_message": str(e),
                        "completed_at": datetime.now(timezone.utc)
                    }
                }
            )
            return False

    async def execute_scheduled_report(
        self,
        schedule_id: str,
        user_id: str
    ) -> Dict[str, Any]:
        """
        Execute a scheduled report and generate the export file.

        Args:
            schedule_id: Scheduled report ID
            user_id: User ID for ownership verification

        Returns:
            Dictionary with execution results
        """
        from app.models.analyst import (
            ScheduledReport, ReportType, ExportFormat, DeliveryMethod
        )
        import uuid

        try:
            # Get schedule from database
            schedule_doc = await self.db.scheduled_reports.find_one({
                "schedule_id": schedule_id,
                "user_id": user_id
            })

            if not schedule_doc:
                return {
                    "success": False,
                    "error": "Scheduled report not found"
                }

            schedule = ScheduledReport.from_mongo(schedule_doc)

            # Map report type to export type
            report_type_to_export = {
                ReportType.DAILY_SUMMARY: ExportType.ANALYTICS,
                ReportType.WEEKLY_ANALYTICS: ExportType.ANALYTICS,
                ReportType.MONTHLY_OVERVIEW: ExportType.ANALYTICS,
                ReportType.HAZARD_REPORT: ExportType.REPORTS,
                ReportType.CUSTOM: ExportType.CUSTOM,
            }
            export_type = report_type_to_export.get(schedule.report_type, ExportType.ANALYTICS)

            # Determine date range based on report type
            if schedule.report_type == ReportType.DAILY_SUMMARY:
                date_range = {"relative": "7days"}
            elif schedule.report_type == ReportType.WEEKLY_ANALYTICS:
                date_range = {"relative": "7days"}
            elif schedule.report_type == ReportType.MONTHLY_OVERVIEW:
                date_range = {"relative": "30days"}
            else:
                date_range = {"relative": "30days"}

            # Fetch data
            data = await self.get_export_data(
                export_type=export_type,
                query_config=schedule.query_config,
                date_range=date_range,
                columns=None
            )

            # Generate file based on format
            export_format = schedule.export_format
            if export_format == ExportFormat.CSV:
                content = await self.generate_csv(data)
                extension = "csv"
                content_type = "text/csv"
            elif export_format == ExportFormat.EXCEL:
                content = await self.generate_excel(data, export_type)
                extension = "xlsx"
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:  # PDF
                title = f"{schedule.name} - {schedule.report_type.value.replace('_', ' ').title()}"
                content = await self.generate_pdf(data, export_type, title)
                extension = "pdf"
                content_type = "application/pdf"

            # Generate file name
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            file_name = f"scheduled_{schedule_id}_{timestamp}.{extension}"
            file_path = EXPORT_DIR / file_name

            # Save file
            with open(file_path, 'wb') as f:
                f.write(content)

            # Update schedule last run info
            now = datetime.now(timezone.utc)
            await self.db.scheduled_reports.update_one(
                {"schedule_id": schedule_id},
                {
                    "$set": {
                        "last_run": now,
                        "last_status": "success",
                        "last_error": None,
                        "updated_at": now
                    }
                }
            )

            # Create an export job record for download access
            job_id = f"EXP-{uuid.uuid4().hex[:8].upper()}"
            export_job = {
                "job_id": job_id,
                "user_id": user_id,
                "export_type": export_type.value,
                "export_format": export_format.value,
                "query_config": schedule.query_config,
                "date_range": date_range,
                "columns": None,
                "status": ExportStatus.COMPLETED.value,
                "progress": 100,
                "file_path": str(file_path),
                "file_name": file_name,
                "file_size": len(content),
                "record_count": len(data) if isinstance(data, list) else 1,
                "created_at": now,
                "started_at": now,
                "completed_at": now,
                "scheduled_report_id": schedule_id
            }
            await self.db.export_jobs.insert_one(export_job)

            logger.info(f"Scheduled report {schedule_id} executed successfully: {file_name}")

            return {
                "success": True,
                "job_id": job_id,
                "file_name": file_name,
                "file_path": str(file_path),
                "file_size": len(content),
                "record_count": len(data) if isinstance(data, list) else 1,
                "export_format": extension,
                "message": f"Report generated successfully. Download available at /analyst/export/{job_id}/download"
            }

        except Exception as e:
            logger.error(f"Error executing scheduled report {schedule_id}: {e}")

            # Update schedule with error
            await self.db.scheduled_reports.update_one(
                {"schedule_id": schedule_id},
                {
                    "$set": {
                        "last_run": datetime.now(timezone.utc),
                        "last_status": "failed",
                        "last_error": str(e),
                        "updated_at": datetime.now(timezone.utc)
                    }
                }
            )

            return {
                "success": False,
                "error": str(e)
            }


# Singleton instance helper
_export_service: Optional[ExportService] = None


def get_export_service(db: AsyncIOMotorDatabase) -> ExportService:
    """Get or create export service instance."""
    global _export_service
    if _export_service is None or _export_service.db != db:
        _export_service = ExportService(db)
    return _export_service


__all__ = ['ExportService', 'get_export_service']
